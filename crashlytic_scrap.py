# scrapping

import undetected_chromedriver as uc
from selenium import webdriver
from openpyxl import load_workbook
import time

#tambahan

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import re

#end

# non headless
options = uc.ChromeOptions()
options.add_argument("--user-data-dir=/Users/ruangguru/Library/Application Support/Google/Chrome/Profile 1") # see notes
driver = uc.Chrome(options=options, use_subprocess=True)
options.add_experimental_option("detach", True)

# # headless
# options = uc.ChromeOptions()
# options.add_argument("--user-data-dir=/Users/ruangguru/Library/Application Support/Google/Chrome/Profile 1")
# options.add_argument('--headless')
# driver = uc.Chrome(options=options, use_subprocess=True)
# options.add_experimental_option("detach", True)

# data store
# moduleId = 'https://teman.sirogu.com/products/14?m=416' # untuk store data URL Module
# sectionID = 'https://teman.sirogu.com/products/93?s=10836' # untuk store data URL Section
# subsectionId = 'https://teman.sirogu.com/products/14?s=11253' # untuk store data URL Sub Section


wb = load_workbook("/Users/ruangguru/Documents/python-automation-test/src/automation.xlsx")

# Select the target sheet
sheet = wb['Crashlytics']

# Clear existing data in the sheet starting from row A2
sheet.delete_rows(2, sheet.max_row)



driver.get("https://play.google.com/console/u/0/developers/6784627491864846493/app/4972886570322244647/vitals/crashes?errorType=CRASH&versionCode=2021066400")
driver.implicitly_wait(10)


container = driver.find_elements(By.XPATH, "//div/root/console-chrome[@helpcentercontext='Play Console']/div/div/div[1]")


print('The list of issues :')

for report in container:
    issuesName = report.find_elements(By.XPATH, "//div[@aria-label='Crashes clusters']//div//ess-cell[1]//console-table-text-cell")
    issuesLink = report.find_elements(By.XPATH, '//div[1]/root[1]/console-chrome[1]/div[1]/div[1]/div[1]/div[1]/div[1]/page-router-outlet[1]/page-wrapper[1]/div[1]/vitals-crashes-page[1]/div[1]/div[1]/console-block-1-column[1]/div[1]/div[1]/crashes-clusters-table[1]/console-table[1]/div[1]/div[1]/ess-table[1]/ess-particle-table[1]/div[1]/div[1]/div[2]/div/ess-cell/console-table-main-action-cell/a')
    for issueNames, issuesURL in zip (issuesName, issuesLink):
        issueList = issueNames.text
        sheet.cell(row=sheet.max_row + 1, column=1, value=issueList)

        urlList = issuesURL.get_attribute('href')
        sheet.cell(row=sheet.max_row + 0, column=2, value=urlList)

        print(issueList)
        print(urlList)
   

wb.save('/Users/ruangguru/Documents/python-automation-test/src/automation.xlsx')
driver.quit()

