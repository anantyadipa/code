# scrapping

import undetected_chromedriver as uc
from selenium import webdriver
from openpyxl import load_workbook
import time

#tambahan

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import re

#end

# non headless
options = uc.ChromeOptions()
options.add_argument("--user-data-dir=/Users/ruangguru/Library/Application Support/Google/Chrome/Profile 1") # see notes
driver = uc.Chrome(options=options, use_subprocess=True)
options.add_experimental_option("detach", True)

# # headless
# options = uc.ChromeOptions()
# options.add_argument("--user-data-dir=/Users/ruangguru/Library/Application Support/Google/Chrome/Profile 1")
# options.add_argument('--headless')
# driver = uc.Chrome(options=options, use_subprocess=True)
# options.add_experimental_option("detach", True)


# wb = load_workbook("/Users/ruangguru/Documents/python-automation-test/src/automation.xlsx")

# # Select the target sheet
# sheet = wb['Crashlytics']

# # Clear existing data in the sheet starting from row A2
# sheet.delete_rows(2, sheet.max_row)



driver.get("https://console.firebase.google.com/u/0/project/firechat-6bf65/crashlytics/app/android:com.ruangguru.livestudents/issues?time=1690070400000:1690329599000&state=open&tag=all&sort=eventCount&types=crash&versions=6.64.1%20(2021066401)")
containerList = driver.find_elements(By.XPATH, "//fireconsole-app-root[1]/fireconsole-home[1]/main[1]/fire-router-outlet[1]/crashlytics-index[1]/c9s-issues[1]/c9s-issues-index[1]/div[1]/div[1]/div[1]/c9s-issues-table[1]")
driver.implicitly_wait(10)





print('The list of issues :')


# for report in container:
#     issuesName = report.find_elements(By.XPATH, "//div[@aria-label='Crashes clusters']//div//ess-cell[1]//console-table-text-cell")
#     issuesLink = report.find_elements(By.XPATH, '//div[1]/root[1]/console-chrome[1]/div[1]/div[1]/div[1]/div[1]/div[1]/page-router-outlet[1]/page-wrapper[1]/div[1]/vitals-crashes-page[1]/div[1]/div[1]/console-block-1-column[1]/div[1]/div[1]/crashes-clusters-table[1]/console-table[1]/div[1]/div[1]/ess-table[1]/ess-particle-table[1]/div[1]/div[1]/div[2]/div/ess-cell/console-table-main-action-cell/a')
#     for issueNames, issuesURL in zip (issuesName, issuesLink):
#         issueList = issueNames.text
#         sheet.cell(row=sheet.max_row + 1, column=1, value=issueList)

#         urlList = issuesURL.get_attribute('href')
#         sheet.cell(row=sheet.max_row + 0, column=2, value=urlList)
        
#         print(issueList)
#         print(urlList)
   


# issueURL = driver.find_elements(By.XPATH, "//fireconsole-app-root/fireconsole-home/main/fire-router-outlet/crashlytics-index/c9s-issues/c9s-issues-index/div/div/div/c9s-issues-table/mat-card/div[2]/table/tbody/tr[1]/td[2]/a")
issuesName = driver.find_elements(By.XPATH, "//fireconsole-app-root[1]/fireconsole-home[1]/main[1]/fire-router-outlet[1]/crashlytics-index[1]/c9s-issues[1]/c9s-issues-index[1]/div[1]/div[1]/div[1]/c9s-issues-table[1]/mat-card[1]/div[2]/table[1]/tbody[1]/tr/td/a/issue-caption-table-cell/div/div[1]/span")
eventsCount = driver.find_elements(By.XPATH, "//fireconsole-app-root[1]/fireconsole-home[1]/main[1]/fire-router-outlet[1]/crashlytics-index[1]/c9s-issues[1]/c9s-issues-index[1]/div[1]/div[1]/div[1]/c9s-issues-table[1]/mat-card[1]/div[2]/table[1]/tbody[1]/tr/td[4]")
usersCount = driver.find_elements(By.XPATH, "//fireconsole-app-root[1]/fireconsole-home[1]/main[1]/fire-router-outlet[1]/crashlytics-index[1]/c9s-issues[1]/c9s-issues-index[1]/div[1]/div[1]/div[1]/c9s-issues-table[1]/mat-card[1]/div[2]/table[1]/tbody[1]/tr/td[5]")

for issueNames, countEvents, countUsers in zip (issuesName, eventsCount, usersCount):
    issueList = issueNames.text
    events = countEvents.text
    users = countUsers.text
    final = print(f'{issueList} | Events: {events} / affected: {users} users')
    print('\n')



# wb.save('/Users/ruangguru/Documents/python-automation-test/src/automation.xlsx')
driver.quit()

