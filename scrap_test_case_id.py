# scrapping

import undetected_chromedriver as uc
from selenium import webdriver
from openpyxl import load_workbook
import time

#tambahan

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import re

#end


# non headless
options = uc.ChromeOptions()
options.add_argument("--user-data-dir=/Users/ruangguru/Library/Application Support/Google/Chrome/Profile 1") # see notes
driver = uc.Chrome(options=options, use_subprocess=True)
options.add_experimental_option("detach", True)

# # headless
# options = uc.ChromeOptions()
# options.add_argument("--user-data-dir=/Users/ruangguru/Library/Application Support/Google/Chrome/Profile 1")
# options.add_argument('--headless')
# driver = uc.Chrome(options=options, use_subprocess=True)
# options.add_experimental_option("detach", True)

# data store
moduleId = 'https://teman.sirogu.com/products/14?m=416' # untuk store data URL Module
sectionID = 'https://teman.sirogu.com/products/93?s=10836' # untuk store data URL Section
subsectionId = 'https://teman.sirogu.com/products/93?s=10839' # untuk store data URL Sub Section


wb = load_workbook('/Users/ruangguru/Documents/python-automation-test/src/test QA automation test case.xlsx')

# Select the target sheet
sheet = wb['Store Id']

# Clear existing data in the sheet starting from row A2
sheet.delete_rows(2, sheet.max_row)

driver.get(subsectionId)
driver.implicitly_wait(10)


WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//p[text()= 'Sub Section 1']")))
# container = driver.find_elements(By.XPATH, "//div[@role='region' and @id='accordion-panel-25']")
testcase_id = driver.find_elements(By.XPATH, "//div[@role='region']/div//a//span")
testcase_title = driver.find_elements(By.XPATH, "//div[@role='region']/div//a//p") 

print('The list of Test Case IDs :')

# # ini untuk scrap TC ID doang
# for id_element in testcase_id:
#     testcase_id_text = re.sub(r"TC-", "", id_element.text)
#     sheet.cell(row=sheet.max_row + 1, column=1, value=testcase_id_text)
#     print(testcase_id_text)

# # ini untuk scrap TC Title doang
# for title_element in testcase_title:
#     testcase_title_text = re.sub(r"Add case", "", title_element.text)
#     # testcase_title_text = title_element.text
#     print(testcase_title_text)

# ini untuk scrap TC ID dan TC Title
for id_element, title_element in zip(testcase_id, testcase_title):
    testcase_id_text = re.sub(r"TC-", "", id_element.text)
    sheet.cell(row=sheet.max_row + 1, column=1, value=testcase_id_text)
    testcase_title_text = title_element.text
    sheet.cell(row=sheet.max_row + 0, column=2, value=testcase_title_text)
    print(f'{testcase_id_text} - {testcase_title_text}')
  
wb.save('/Users/ruangguru/Documents/python-automation-test/src/test QA automation test case.xlsx')
driver.quit()

